from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from Library import configreder
from Base import initiateDruver
from Pages import registrationpge
import pytest
import openpyxl


def dataGenerator():
    wb=openpyxl.load_workbook("D:\Book1.xlsx")
    sh=wb['Sheet1']
    r=sh.max_row
    li=[]
    li1=[]
    for i in range(1,r+1):
        li1=[]
        un=sh.cell(i,1)
        up=sh.cell(i,2)
        li1.insert(0,un.value)
        li1.insert(1,up.value)
        li.insert(i-1,li1)

    print(li)
    return li

@pytest.mark.parametrize('data',dataGenerator())
def test_validregis(data):
    driver = initiateDruver.start_br()
    regis=registrationpge.registion(driver)
    regis.enterusername(data[0])
    regis.emil(data[1])
